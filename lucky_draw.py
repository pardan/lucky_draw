from openpyxl import load_workbook
import random
import schedule
import time


def main():
    wb = load_workbook('lucky_draw_py\Data Dummy Pengundian2.xlsx')
    sheets = wb.sheetnames
    ws = wb[sheets[0]]

    column = 'A'
    row_range = list(range(2, 66341))

    rand_cell = "{column}{row}".format(column = column, row = random.choice(row_range))

    print(ws[rand_cell].value)


if __name__ == '__main__':
    main()
    schedule.every(0).seconds.do(main)
    while True:
        schedule.run_pending()
        time.sleep(0)
